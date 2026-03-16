import pandas as pd
import json
import os
from datetime import datetime
import pytz

from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters, ConversationHandler

from openpyxl import load_workbook
import shutil
from openpyxl.styles import PatternFill

TOKEN = "8668806663:AAGxThCG0tfr5DY_Pw-E1gMjNQSgePLydfY"

ADMIN_IDS = [65604662, 34600346, 147411723]

# -----------------------------
# BACA FAIL EXCEL
# -----------------------------

files = [
"TINGKATAN 1.xlsx",
"TINGKATAN 2.xlsx",
"TINGKATAN 3.xlsx",
"TINGKATAN 4.xlsx",
"TINGKATAN 5.xlsx",
"STAM.xlsx"
]

kelas_data = {}

for file in files:

    xl = pd.ExcelFile(file)

    for sheet in xl.sheet_names:

        df = xl.parse(sheet)

        nama = df.iloc[:,0].dropna().astype(str).tolist()

        kelas_data[sheet] = nama

kelas_list = list(kelas_data.keys())

# -----------------------------
# GURU KELAS
# -----------------------------

guru_kelas = {

"LULU": "Pn Rohani",
"MARJAN": "En Sobur",

"5K": "En Ahmad Shahir",
"5A": "Pn Hjh Norhuda",
"5M": "Pn Najibah",

"4K": "En Mohd Nursyawal",
"4A": "En Sharifuddin",
"4M": "En Zaifuan",

"3K": "Cik Najwa",
"3A": "Pn Dayang",
"3M": "En Mohd Johari",

"2K": "En Mohmad Rozi",
"2A": "Pn Salmiah",
"2M": "Cik Norshafikah",

"1K": "Pn Haniza",
"1A": "Pn Zulaikah",
"1M": "Pn Hanizan"

}

# -----------------------------
# DATA SISTEM
# -----------------------------

kelas_dipilih = {}
tidak_hadir = {}
kelas_selesai = set()
pelajar_dipilih = {}

DATA_FILE = "data_kehadiran.json"

# -----------------------------
# SENARAI SEBAB
# -----------------------------

sebab_data = {

"Aktiviti Luar Sekolah": ["Aktiviti luar sekolah"],
"Digantung Sekolah": ["Digantung sekolah"],
"Ponteng": ["Ponteng sekolah"],
"Bencana Alam": ["Bencana alam"],
"Ancaman Keselamatan": ["Ancaman keselamatan"],
"Masalah Peribadi": ["Masalah peribadi"],
"Masalah Kesihatan": ["Masalah kesihatan"],
"Masalah Keluarga": ["Masalah keluarga"],
"Kebenaran Pengetua / GB": ["Kebenaran pengetua"]

}

# -----------------------------
# STATE
# -----------------------------

PILIH_KELAS, PILIH_PELAJAR, PILIH_SEBAB = range(3)

# -----------------------------
# SIMPAN DATA
# -----------------------------

def simpan_data():

    data = {
        "tarikh": masa_malaysia().strftime("%Y-%m-%d"),
        "kelas_selesai": list(kelas_selesai),
        "tidak_hadir": tidak_hadir
    }

    with open(DATA_FILE, "w") as f:
        json.dump(data, f)

def simpan_rekod_harian():

    hari, tarikh = hari_tarikh()

    data_kelas = {}

    jumlah_sekolah = 0
    hadir_sekolah = 0

    for kelas in kelas_list:

        jumlah = len(kelas_data[kelas])
        tidak = len(tidak_hadir.get(kelas, []))
        hadir = jumlah - tidak

        if jumlah > 0:
            peratus = (hadir / jumlah) * 100
        else:
            peratus = 0

        data_kelas[kelas] = {
            "jumlah": jumlah,
            "hadir": hadir,
            "tidak": tidak,
            "peratus": round(peratus,2)
        }

        jumlah_sekolah += jumlah
        hadir_sekolah += hadir

    tidak_sekolah = jumlah_sekolah - hadir_sekolah

    if jumlah_sekolah > 0:
        peratus_sekolah = (hadir_sekolah / jumlah_sekolah) * 100
    else:
        peratus_sekolah = 0

    rekod = {
        "tarikh": tarikh,
        "hari": hari,
        "kelas": data_kelas,
        "peratus_sekolah": round(peratus_sekolah,2)
    }

    file = "rekod_kehadiran.json"

    if os.path.exists(file):

        with open(file) as f:
            data = json.load(f)

    else:
        data = []

    data.append(rekod)

    with open(file,"w") as f:
        json.dump(data,f,indent=4)

# -----------------------------
# LOAD DATA
# -----------------------------

def load_data():

    global kelas_selesai, tidak_hadir

    if os.path.exists(DATA_FILE):

        with open(DATA_FILE) as f:

            data = json.load(f)

            tarikh_simpan = data.get("tarikh")
            tarikh_hari_ini = masa_malaysia().strftime("%Y-%m-%d")

            if tarikh_simpan == tarikh_hari_ini:

                kelas_selesai = set(data.get("kelas_selesai",[]))
                tidak_hadir = data.get("tidak_hadir",{})

            else:

                kelas_selesai = set()
                tidak_hadir = {}

# -----------------------------
# TARIKH
# -----------------------------

def masa_malaysia():

    tz = pytz.timezone("Asia/Kuala_Lumpur")

    return datetime.now(tz)

def hari_tarikh():

    now = masa_malaysia()

    hari_en = now.strftime("%A")

    hari_melayu = {
        "Monday": "Isnin",
        "Tuesday": "Selasa",
        "Wednesday": "Rabu",
        "Thursday": "Khamis",
        "Friday": "Jumaat",
        "Saturday": "Sabtu",
        "Sunday": "Ahad"
    }

    hari = hari_melayu.get(hari_en, hari_en)

    tarikh = now.strftime("%d-%m-%Y")

    return hari, tarikh

# -----------------------------
# CARI COLUMN TARIKH DALAM EXCEL
# -----------------------------

def cari_column_tarikh(sheet, tarikh):

    tarikh = str(tarikh)

    for col in range(1, sheet.max_column + 1):

        cell = sheet.cell(row=4, column=col).value

        if cell is None:
            continue

        if tarikh in str(cell):
            return col

    return None

# -----------------------------
# START
# -----------------------------

async def start(update:Update,context:ContextTypes.DEFAULT_TYPE):

    hari,tarikh = hari_tarikh()

    keyboard = [[k] for k in kelas_list]

    reply_markup = ReplyKeyboardMarkup(keyboard,resize_keyboard=True)

    await update.message.reply_text(
        f"SISTEM KEHADIRAN\n\n"
        f"Hari: {hari}\n"
        f"Tarikh: {tarikh}\n\n"
        f"Pilih kelas:",
        reply_markup=reply_markup
    )

    return PILIH_KELAS

# -----------------------------
# PILIH KELAS
# -----------------------------

async def pilih_kelas(update:Update,context:ContextTypes.DEFAULT_TYPE):

    teks = update.message.text
    chat_id = update.message.chat_id

    if teks not in kelas_list:
        return PILIH_KELAS

    if teks in kelas_selesai:

        guru = guru_kelas.get(teks,"-")

        await update.message.reply_text(
            f"Kelas {teks} telah direkod.\n"
            f"Guru kelas: {guru}\n\n"
            f"Jika ada kesilapan, admin boleh reset."
        )

        return ConversationHandler.END

    if teks in kelas_selesai:

        guru = guru_kelas.get(teks,"-")

        await update.message.reply_text(
            f"Kelas {teks} telah direkod.\n"
            f"Guru kelas: {guru}"
        )

        return ConversationHandler.END

    kelas_dipilih[chat_id] = teks
    tidak_hadir.setdefault(teks,[])

    pelajar = kelas_data[teks]

    keyboard = [[p] for p in pelajar]
    keyboard.append(["SELESAI"])
    keyboard.append(["KEMBALI"])

    reply_markup = ReplyKeyboardMarkup(keyboard,resize_keyboard=True)

    await update.message.reply_text(
        f"Kelas {teks}\nKlik nama pelajar tidak hadir.",
        reply_markup=reply_markup
    )

    return PILIH_PELAJAR

# -----------------------------
# PILIH PELAJAR
# -----------------------------

async def pilih_pelajar(update:Update,context:ContextTypes.DEFAULT_TYPE):

    chat_id = update.message.chat_id
    teks = update.message.text

    if teks == "KEMBALI":

        keyboard = [[k] for k in kelas_list]

        reply_markup = ReplyKeyboardMarkup(keyboard,resize_keyboard=True)

        await update.message.reply_text(
            "Pilih kelas semula:",
            reply_markup=reply_markup
        )

        return PILIH_KELAS

    kelas = kelas_dipilih[chat_id]

    if teks == "SELESAI":

        kelas_selesai.add(kelas)
        simpan_data()

        if len(kelas_selesai) == len(kelas_list):
            simpan_rekod_harian()

        jumlah = len(kelas_data[kelas])
        tidak = len(tidak_hadir[kelas])
        hadir = jumlah - tidak

        hari, tarikh = hari_tarikh()

        teks_laporan = f"KELAS {kelas} SELESAI\n"
        teks_laporan += f"Hari: {hari}\n"
        teks_laporan += f"Tarikh: {tarikh}\n"
        teks_laporan += f"Jumlah: {jumlah}\n"
        teks_laporan += f"Hadir: {hadir}\n"
        teks_laporan += f"Tidak Hadir: {tidak}\n"

        if tidak > 0:

            teks_laporan += "\nSenarai Tidak Hadir:\n"

            for i, pelajar in enumerate(tidak_hadir[kelas], start=1):

                nama = pelajar["nama"]
                sebab = pelajar["sebab"]

                teks_laporan += f"{i}. {nama} ({sebab})\n"

        else:

            teks_laporan += "\nTiada pelajar tidak hadir\n"

        await update.message.reply_text(teks_laporan)

        return ConversationHandler.END

    pelajar_dipilih[chat_id] = teks

    keyboard = [[k] for k in sebab_data.keys()]
    keyboard.append(["KEMBALI"])

    reply_markup = ReplyKeyboardMarkup(keyboard,resize_keyboard=True)

    await update.message.reply_text(
        f"Pilih sebab tidak hadir bagi {teks}:",
        reply_markup=reply_markup
    )

    return PILIH_SEBAB

# -----------------------------
# PILIH SEBAB
# -----------------------------

async def pilih_sebab(update:Update,context:ContextTypes.DEFAULT_TYPE):

    chat_id = update.message.chat_id
    teks = update.message.text

    kelas = kelas_dipilih[chat_id]
    pelajar = pelajar_dipilih[chat_id]

    if teks == "KEMBALI":
        return PILIH_PELAJAR

    tidak_hadir.setdefault(kelas, [])

    # semak jika pelajar sudah direkod
    sudah_ada = False

    for p in tidak_hadir[kelas]:

        if p["nama"] == pelajar:

            sudah_ada = True
            break

    if not sudah_ada:

        tidak_hadir[kelas].append({
            "nama": pelajar,
            "sebab": teks
        })

    simpan_data()

    await update.message.reply_text(
        f"{pelajar} direkod tidak hadir\nSebab: {teks}"
    )

    pelajar_list = kelas_data[kelas]

    keyboard = [[p] for p in pelajar_list]
    keyboard.append(["SELESAI"])
    keyboard.append(["KEMBALI"])

    reply_markup = ReplyKeyboardMarkup(keyboard,resize_keyboard=True)

    await update.message.reply_text(
        "Pilih pelajar lain atau tekan SELESAI:",
        reply_markup=reply_markup
    )

    return PILIH_PELAJAR

# -----------------------------
# STATUS KEHADIRAN
# -----------------------------

async def status(update: Update, context: ContextTypes.DEFAULT_TYPE):

    hari, tarikh = hari_tarikh()

    teks = "STATUS KEHADIRAN KELAS\n"
    teks += f"TARIKH: {tarikh}\n"
    teks += f"HARI: {hari}\n\n"

    for k in kelas_list:

        guru = guru_kelas.get(k, "-")

        if k in kelas_selesai:
            teks += f"{k} - {guru} ✅\n"
        else:
            teks += f"{k} - {guru} ❌\n"

    await update.message.reply_text(teks)


# -----------------------------
# RESET KELAS (ADMIN)
# -----------------------------

async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if update.message.from_user.id not in ADMIN_IDS:

        await update.message.reply_text("Hanya admin boleh reset.")
        return

    if not context.args:

        await update.message.reply_text("Contoh: /reset 4A")
        return

    kelas = context.args[0].upper()

    if kelas in kelas_selesai:

        kelas_selesai.remove(kelas)

        tidak_hadir[kelas] = []

        simpan_data()

        await update.message.reply_text(
            f"Kelas {kelas} telah direset."
        )

    else:

        await update.message.reply_text(
            f"Kelas {kelas} belum direkod."
        )

# -----------------------------
# LAPORAN SEMASA
# -----------------------------

async def laporan_semasa(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if update.message.from_user.id not in ADMIN_IDS:

        await update.message.reply_text(
            "Command ini hanya untuk admin."
        )

        return

    hari, tarikh = hari_tarikh()

    teks = "LAPORAN KEHADIRAN SEMASA\n\n"

    for kelas in kelas_list:

        teks += f"KELAS {kelas}\n"
        teks += f"Hari: {hari}\n"
        teks += f"Tarikh: {tarikh}\n"

        if kelas not in kelas_selesai:

            teks += "Status: Belum isi kehadiran\n\n"
            continue

        jumlah = len(kelas_data[kelas])
        tidak = len(tidak_hadir.get(kelas, []))
        hadir = jumlah - tidak
        peratus = hadir / jumlah * 100

        teks += f"Jumlah: {jumlah}\n"
        teks += f"Hadir: {hadir}\n"
        teks += f"Tidak Hadir: {tidak}\n"
        teks += f"Peratus: {peratus:.2f}%\n\n"

    await update.message.reply_text(teks)


# -----------------------------
# LAPORAN PENUH
# -----------------------------

async def laporan_penuh(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if update.message.from_user.id not in ADMIN_IDS:

        await update.message.reply_text(
            "Command ini hanya untuk admin."
        )

        return

    if len(kelas_selesai) < len(kelas_list):

        await update.message.reply_text(
            "Masih ada kelas belum isi kehadiran.\n"
            "Gunakan /status untuk semak."
        )

        return    

    hari, tarikh = hari_tarikh()

    jumlah_sekolah = 0
    hadir_sekolah = 0

    for kelas in kelas_list:

        jumlah = len(kelas_data[kelas])
        tidak = len(tidak_hadir.get(kelas, []))
        hadir = jumlah - tidak

        jumlah_sekolah += jumlah
        hadir_sekolah += hadir

    tidak_hadir_sekolah = jumlah_sekolah - hadir_sekolah
    peratus = hadir_sekolah / jumlah_sekolah * 100

    teks = "LAPORAN KEHADIRAN SEKOLAH\n\n"
    teks += f"Hari: {hari}\n"
    teks += f"Tarikh: {tarikh}\n\n"
    teks += f"Jumlah Pelajar Sekolah: {jumlah_sekolah}\n"
    teks += f"Hadir: {hadir_sekolah}\n"
    teks += f"Tidak Hadir: {tidak_hadir_sekolah}\n"
    teks += f"Peratus Kehadiran: {peratus:.2f}%"

    await update.message.reply_text(teks)


# -----------------------------
# LAPORAN KELAS
# -----------------------------

async def laporan_kelas(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if update.message.from_user.id not in ADMIN_IDS:

        await update.message.reply_text(
            "Command ini hanya untuk admin."
        )

        return

    hari, tarikh = hari_tarikh()

    teks = "LAPORAN KETIDAKHADIRAN SETIAP KELAS\n\n"

    for kelas in kelas_list:

        teks += f"KELAS {kelas}\n"
        teks += f"Hari: {hari}\n"
        teks += f"Tarikh: {tarikh}\n"

        if kelas not in kelas_selesai:

            teks += "Belum isi kehadiran\n\n"
            continue

        tidak = tidak_hadir.get(kelas, [])

        if len(tidak) == 0:

            teks += "Tiada pelajar tidak hadir\n\n"

        else:

            teks += "Nama yang tidak hadir:\n"

            for i, pelajar in enumerate(tidak, start=1):

                nama = pelajar["nama"]
                sebab = pelajar["sebab"]

            teks += f"{i}. {nama} ({sebab})\n"

            teks += "\n"

    await update.message.reply_text(teks)

async def excel_harian(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if update.message.from_user.id not in ADMIN_IDS:
        await update.message.reply_text("Command ini hanya untuk admin.")
        return

    hari, tarikh = hari_tarikh()

    fail = "data_kehadiran.xlsx"

    if not os.path.exists(fail):
        shutil.copy("template_kehadiran.xlsx", fail)

    wb = load_workbook(fail)

    sheet = wb["PERATUS HARIANMINGGUAN"]

    column = cari_column_tarikh(sheet, tarikh)

    if column is None:

        await update.message.reply_text("Tarikh tidak dijumpai dalam template Excel.")
        return

    row = 19

    for kelas in kelas_list:

        jumlah = len(kelas_data[kelas])
        tidak = len(tidak_hadir.get(kelas, []))
        hadir = jumlah - tidak

        if jumlah > 0:
            peratus = round((hadir / jumlah) * 100, 2)
        else:
            peratus = 0

        sheet.cell(row=row, column=column).value = peratus

        row += 1


# -----------------------------
# WARNA MERAH JIKA TIADA DATA
# -----------------------------

    from openpyxl.styles import PatternFill

    merah = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    for col in range(2, sheet.max_column + 1):

        cell = sheet.cell(row=19, column=col)

        if cell.value is None:

            for r in range(19, 19 + len(kelas_list)):

                sheet.cell(row=r, column=col).fill = merah


    wb.save(fail)
    wb.close()

    await update.message.reply_document(document=open(fail, "rb"))

async def excel_mingguan(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if update.message.from_user.id not in ADMIN_IDS:
        return

    fail = "data_kehadiran.xlsx"

    if not os.path.exists(fail):

        await update.message.reply_text("Tiada rekod Excel.")
        return

    await update.message.reply_document(document=open(fail,"rb"))

# -----------------------------
# RESET SEMUA KEHADIRAN (ADMIN)
# -----------------------------

async def reset_semua(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if update.message.from_user.id not in ADMIN_IDS:

        await update.message.reply_text("Hanya admin boleh reset semua kehadiran.")
        return

    kelas_selesai.clear()
    tidak_hadir.clear()

    simpan_data()

    await update.message.reply_text(
        "SEMUA KEHADIRAN HARI INI TELAH DIRESET.\nGuru boleh isi semula kehadiran."
    )

async def laporan_mingguan(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if update.message.from_user.id not in ADMIN_IDS:
        return

    file = "rekod_kehadiran.json"

    if not os.path.exists(file):
        await update.message.reply_text("Tiada rekod kehadiran.")
        return

    with open(file) as f:
        data = json.load(f)

    minggu = data[-5:]

    teks = "LAPORAN KEHADIRAN MINGGUAN\n\n"

    jumlah_peratus = 0

    kelas_stat = {k:[] for k in kelas_list}

    teks += "SENARAI KEHADIRAN HARIAN\n\n"

    for d in minggu:

        tarikh = d["tarikh"]
        hari = d["hari"]
        peratus = d["peratus_sekolah"]

        teks += f"{tarikh} ({hari}) : {peratus:.2f}%\n"

        jumlah_peratus += peratus

        for kelas in kelas_list:
            kelas_stat[kelas].append(d["kelas"][kelas]["peratus"])

    purata_sekolah = jumlah_peratus / len(minggu)

    teks += "\n"
    teks += f"Purata Kehadiran Sekolah : {purata_sekolah:.2f}%\n\n"

    teks += "PERATUS KEHADIRAN KELAS\n\n"

    for kelas in kelas_list:

        purata = sum(kelas_stat[kelas]) / len(kelas_stat[kelas])

        teks += f"{kelas} : {purata:.2f}%\n"

    await update.message.reply_text(teks)

import calendar
from datetime import datetime

async def laporan_bulanan(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if update.message.from_user.id not in ADMIN_IDS:
        return

    file = "rekod_kehadiran.json"

    if not os.path.exists(file):
        await update.message.reply_text("Tiada rekod kehadiran.")
        return

    with open(file) as f:
        data = json.load(f)

    sekarang = datetime.now()
    tahun = sekarang.year
    bulan = sekarang.month

    nama_bulan = sekarang.strftime("%B").upper()

    jumlah_hari = calendar.monthrange(tahun, bulan)[1]

    rekod_tarikh = {}

    kelas_stat = {k:[] for k in kelas_list}

    for d in data:

        rekod_tarikh[d["tarikh"]] = d["peratus_sekolah"]

        for kelas in kelas_list:
            kelas_stat[kelas].append(d["kelas"][kelas]["peratus"])

    teks = f"LAPORAN KEHADIRAN BULAN {nama_bulan}\n\n"

    jumlah_peratus = 0
    bil_rekod = 0

    teks += "SENARAI KEHADIRAN HARIAN\n\n"

    for hari in range(1, jumlah_hari+1):

        tarikh = f"{hari:02d}-{bulan:02d}-{tahun}"

        if tarikh in rekod_tarikh:

            peratus = rekod_tarikh[tarikh]

            teks += f"{tarikh} : {peratus:.2f}%\n"

            jumlah_peratus += peratus
            bil_rekod += 1

        else:

            teks += f"{tarikh} : Tiada Rekod\n"

    if bil_rekod > 0:
        purata_sekolah = jumlah_peratus / bil_rekod
    else:
        purata_sekolah = 0

    teks += "\n"
    teks += f"Bil Hari Sekolah Direkod : {bil_rekod}\n\n"
    teks += f"Purata Kehadiran Sekolah : {purata_sekolah:.2f}%\n\n"

    teks += "PERATUS KEHADIRAN KELAS\n\n"

    for kelas in kelas_list:

        if len(kelas_stat[kelas]) > 0:

            purata = sum(kelas_stat[kelas]) / len(kelas_stat[kelas])

            teks += f"{kelas} : {purata:.2f}%\n"

    await update.message.reply_text(teks)

# -----------------------------
# TELEGRAM BOT
# -----------------------------

app = ApplicationBuilder().token(TOKEN).build()

app.add_handler(CommandHandler("status", status))
app.add_handler(CommandHandler("reset", reset))
app.add_handler(CommandHandler("resetsemua", reset_semua))

app.add_handler(CommandHandler("laporansemasa", laporan_semasa))
app.add_handler(CommandHandler("laporanpenuh", laporan_penuh))
app.add_handler(CommandHandler("laporankelas", laporan_kelas))

app.add_handler(CommandHandler("excelharian", excel_harian))
app.add_handler(CommandHandler("excelmingguan", excel_mingguan))

app.add_handler(CommandHandler("laporanmingguan", laporan_mingguan))
app.add_handler(CommandHandler("laporanbulanan", laporan_bulanan))

conv_handler = ConversationHandler(

    entry_points=[CommandHandler("start", start)],

    states={

        PILIH_KELAS: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, pilih_kelas)
        ],

        PILIH_PELAJAR: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, pilih_pelajar)
        ],

        PILIH_SEBAB: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, pilih_sebab)
        ]

    },

    fallbacks=[]
)

app.add_handler(conv_handler)

load_data()

print("Bot sedang berjalan...")

app.run_polling()